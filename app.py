That happens because on a lot of NBFWM pages the **province isn’t showing up as “QC/ON/etc” in a spot we reliably captured**, so `address_hint` ends up blank → `province` stays blank.

Best fix: **read the City/Province directly from the main directory page (`/advisor.html`) and use it as a fallback** for each advisor profile URL. That directory page clearly contains lines like **“Montreal, Quebec”**, **“Vancouver, British-Columbia”**, etc., so province becomes rock-solid.

Here’s a **final updated `app.py`** that:

* builds a `profile_url -> city/province` lookup from the directory page
* parses each profile page for email/phone as before
* sets `province` from:

  1. profile address/link, else
  2. the directory lookup

> Replace your current `app.py` with this one and redeploy/reboot your Streamlit app.

```python
# app.py
# NBFWM Advisor Directory Extractor (adds Province reliably via directory fallback)
# Run: streamlit run app.py
#
# requirements.txt:
# streamlit
# requests
# beautifulsoup4
# pandas
# openpyxl

import re
import time
import json
from io import BytesIO
from collections import deque
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
import streamlit as st

# Optional Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


# ----------------------------- UI -----------------------------

st.set_page_config(page_title="NBFWM Advisor Extractor", layout="wide")

st.markdown(
    """
<style>
.block-container { max-width: 1180px; padding-top: 2rem; padding-bottom: 3rem; }
h1, h2, h3 { letter-spacing: -0.02em; }
hr { margin: 1.4rem 0; }
.card {
  border: 1px solid rgba(0,0,0,0.10);
  background: rgba(255,255,255,0.70);
  border-radius: 18px;
  padding: 16px 18px;
  box-shadow: 0 6px 18px rgba(0,0,0,0.06);
}
.small-muted { color: rgba(0,0,0,0.55); font-size: 0.92rem; }
</style>
""",
    unsafe_allow_html=True,
)

st.title("NBFWM Wealth Advisor Directory Extractor")
st.caption("Exports publicly listed advisor info from nbfwm.ca (CSV + optional Excel).")


# ----------------------------- Constants / Regex -----------------------------

DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; InovestorDirectoryExtractor/1.2; +https://inovestor.com)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

BASE_DEFAULT = "https://www.nbfwm.ca"
SEED_PATH_DEFAULT = "/advisor.html"

# NBFWM advisor profile pages look like:
# /advisor/<team-slug>/(our-team|notre-equipe)/<advisor-slug>.html
ADVISOR_HREF_RE = re.compile(r"^/advisor/.+/(our-team|notre-equipe)/.+\.html$", re.I)

EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)
PHONE_RE = re.compile(r"\b(?:1[-\s]?)?\(?\d{3}\)?[-\s]?\d{3}[-\s]?\d{4}\b")

# Canadian provinces/territories abbreviations
PROV_ABBR_RE = re.compile(r"\b(BC|AB|SK|MB|ON|QC|NB|NS|PE|NL|NT|NU|YT)\b", re.I)

# Full names map (include hyphenated forms used on NBFWM like "British-Columbia")
FULL_PROV_MAP = {
    "quebec": "QC", "québec": "QC",
    "ontario": "ON",
    "alberta": "AB",
    "manitoba": "MB",
    "saskatchewan": "SK",
    "british columbia": "BC", "british-columbia": "BC",
    "colombie britannique": "BC", "colombie-britannique": "BC",
    "new brunswick": "NB", "new-brunswick": "NB",
    "nouveau brunswick": "NB", "nouveau-brunswick": "NB",
    "nova scotia": "NS", "nova-scotia": "NS",
    "nouvelle ecosse": "NS", "nouvelle-écosse": "NS", "nouvelle-ecosse": "NS",
    "prince edward island": "PE",
    "ile du prince edouard": "PE", "île du prince édouard": "PE", "île-du-prince-édouard": "PE",
    "newfoundland and labrador": "NL",
    "terre neuve et labrador": "NL", "terre-neuve-et-labrador": "NL",
    "northwest territories": "NT",
    "territoires du nord ouest": "NT", "territoires du nord-ouest": "NT",
    "nunavut": "NU",
    "yukon": "YT",
}

# City, ProvinceName lines from /advisor.html (e.g., "Montreal, Quebec", "Vancouver, British-Columbia")
PROV_NAME_PATTERN = (
    r"(Alberta|British[- ]Columbia|Manitoba|New[- ]Brunswick|Nova[- ]Scotia|Ontario|Quebec|Saskatchewan|"
    r"Prince Edward Island|Newfoundland(?: and Labrador)?|Northwest Territories|Nunavut|Yukon)"
)
CITY_PROV_LINE_RE = re.compile(rf"^\s*([^,\n]+?)\s*,\s*{PROV_NAME_PATTERN}\s*$", re.I)

POSTAL_RE = re.compile(r"\b[ABCEGHJ-NPRSTVXY]\d[ABCEGHJ-NPRSTV-Z]\s?\d[ABCEGHJ-NPRSTV-Z]\d\b", re.I)


# ----------------------------- Helpers -----------------------------

def normalize_phone(p: str) -> str:
    digits = re.sub(r"\D+", "", p or "")
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}"
    return (p or "").strip()


def extract_team_name_from_slug(team_slug: str) -> str:
    if not team_slug:
        return ""
    return team_slug.replace("-", " ").strip().title()


def extract_province(text: str) -> str:
    """Return province code (QC/ON/...) from any text containing full name or abbreviation."""
    if not text:
        return ""
    m = PROV_ABBR_RE.search(text)
    if m:
        return m.group(1).upper()

    t = text.strip().lower()
    # normalize hyphens/apostrophes for matching
    t = t.replace("’", "'")
    t_norm = t.replace("-", " ")

    for k, v in FULL_PROV_MAP.items():
        if k in t or k in t_norm:
            return v
    return ""


def safe_get(session: requests.Session, url: str, delay_s: float, timeout: int = 30) -> str:
    r = session.get(url, headers=DEFAULT_HEADERS, timeout=timeout)
    r.raise_for_status()
    if delay_s and delay_s > 0:
        time.sleep(delay_s)
    return r.text


def same_domain(url: str, base: str) -> bool:
    return urlparse(url).netloc.lower() == urlparse(base).netloc.lower()


def extract_advisor_urls_from_html(html: str, base: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    links = set()

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if ADVISOR_HREF_RE.match(href):
            links.add(urljoin(base, href))

    # fallback regex scan
    for m in re.finditer(r'"/advisor/[^"]+/(?:our-team|notre-equipe)/[^"]+\.html"', html, flags=re.I):
        href = m.group(0).strip('"')
        links.add(urljoin(base, href))

    return sorted(links)


def extract_internal_html_pages_from_html(html: str, page_url: str, base: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    found = set()

    for a in soup.find_all("a", href=True):
        full = urljoin(page_url, a["href"].strip())
        if not same_domain(full, base):
            continue
        p = urlparse(full).path.lower()
        if p.startswith("/advisor/") and p.endswith(".html"):
            found.add(full)

    for m in re.finditer(r'"/advisor/[^"]+\.html"', html, flags=re.I):
        href = m.group(0).strip('"')
        found.add(urljoin(base, href))

    return sorted(found)


def build_directory_location_lookup(directory_html: str, base: str) -> dict:
    """
    Build a mapping: profile_url -> {"city": ..., "province": ...}
    from the /advisor.html page content.
    """
    soup = BeautifulSoup(directory_html, "html.parser")
    lookup = {}

    # Find all advisor profile links; for each, read nearby text from the card/container.
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not ADVISOR_HREF_RE.match(href):
            continue

        profile_url = urljoin(base, href)

        # try to locate the card container
        container = a.find_parent(["li", "article", "section", "div"])
        text = ""
        if container:
            text = container.get_text("\n", strip=True)
        else:
            text = a.parent.get_text("\n", strip=True) if a.parent else ""

        city = ""
        prov_code = ""

        # try to find a line like "Montreal, Quebec"
        for line in [ln.strip() for ln in text.split("\n") if ln.strip()]:
            m = CITY_PROV_LINE_RE.match(line)
            if m:
                city = m.group(1).strip()
                prov_code = extract_province(line)
                break

        if prov_code:
            lookup[profile_url] = {"city": city, "province": prov_code}

    return lookup


def extract_address_hint(soup: BeautifulSoup) -> str:
    """
    Best-effort address: locator link text, JSON-LD schema address, or any line with postal/province.
    """
    # 1) Locator links (often used for the address line)
    for a in soup.find_all("a", href=True):
        href = (a["href"] or "").strip().lower()
        if "locator" in href:
            txt = a.get_text(" ", strip=True)
            if txt:
                return txt

    # 2) JSON-LD schema.org
    for script in soup.find_all("script", type="application/ld+json"):
        raw = script.string or script.get_text(strip=True)
        if not raw:
            continue
        try:
            data = json.loads(raw)
            items = data if isinstance(data, list) else [data]
            for it in items:
                if isinstance(it, dict) and "address" in it:
                    addr = it.get("address")
                    if isinstance(addr, dict):
                        parts = [
                            addr.get("streetAddress", ""),
                            addr.get("addressLocality", ""),
                            addr.get("addressRegion", ""),
                            addr.get("postalCode", ""),
                        ]
                        hint = ", ".join([p for p in parts if p])
                        if hint:
                            return hint
        except Exception:
            # not always strict JSON
            pass

    # 3) Any text line with a Canadian postal code
    lines = [ln.strip() for ln in soup.get_text("\n", strip=True).split("\n") if ln.strip()]
    for ln in lines:
        if POSTAL_RE.search(ln):
            return ln[:220]

    # 4) Any line containing a recognizable province full name
    for ln in lines:
        if extract_province(ln):
            # keep short
            return ln[:220]

    return ""


def parse_advisor_page(html: str, url: str, base: str, dir_lookup: dict) -> dict:
    soup = BeautifulSoup(html, "html.parser")

    # Name
    name = ""
    h1 = soup.find("h1")
    if h1:
        name = h1.get_text(" ", strip=True)
    if not name and soup.title:
        name = soup.title.get_text(" ", strip=True).split("|")[0].strip()

    # Email (prefer mailto)
    email = ""
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.lower().startswith("mailto:"):
            email = href.split(":", 1)[1].split("?", 1)[0].strip()
            break
    if not email:
        all_text = soup.get_text("\n", strip=True)
        emails = sorted(set(EMAIL_RE.findall(all_text)))
        email = emails[0] if emails else ""

    # Phone (prefer tel)
    phones = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.lower().startswith("tel:"):
            phones.append(normalize_phone(href.split(":", 1)[1]))
    if not phones:
        all_text = soup.get_text("\n", strip=True)
        phones = [normalize_phone(p) for p in PHONE_RE.findall(all_text)]
    phones = [p for p in pd.unique(phones) if p]
    phone = " | ".join(list(phones)[:3])

    # Address & province (primary)
    address_hint = extract_address_hint(soup)
    province = extract_province(address_hint)

    # Province fallback from directory lookup (very reliable)
    if not province:
        fallback = dir_lookup.get(url, {})
        province = fallback.get("province", "")

    city = ""
    if dir_lookup.get(url):
        city = dir_lookup[url].get("city", "")

    # Team slug (fallback team name)
    team_name = ""
    try:
        parts = urlparse(url).path.strip("/").split("/")
        team_slug = parts[1] if len(parts) > 1 and parts[0].lower() == "advisor" else ""
        team_name = extract_team_name_from_slug(team_slug)
    except Exception:
        team_name = ""

    return {
        "name": (name or "").strip(),
        "email": (email or "").strip(),
        "phone": (phone or "").strip(),
        "team_name": (team_name or "").strip(),
        "province": (province or "").strip(),
        "city": (city or "").strip(),
        "address_hint": (address_hint or "").strip(),
        "profile_url": url,
    }


def dedupe_rows(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["email_norm"] = df["email"].fillna("").str.lower()
    df["name_norm"] = df["name"].fillna("").str.lower()
    df["phone_norm"] = df["phone"].fillna("").str.replace(r"\s+", "", regex=True)

    has_email = df["email_norm"] != ""
    df_email = df[has_email].drop_duplicates(subset=["email_norm"], keep="first")
    df_no_email = df[~has_email].drop_duplicates(subset=["name_norm", "phone_norm"], keep="first")

    out = pd.concat([df_email, df_no_email], ignore_index=True)
    return out.drop(columns=["email_norm", "name_norm", "phone_norm"], errors="ignore")


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Advisors"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    align = Alignment(vertical="center", wrap_text=False)

    fills = [
        PatternFill("solid", fgColor="FFFFFF"),
        PatternFill("solid", fgColor="F7FBFF"),
    ]

    ws.append(list(df.columns))
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align

    for r in df.itertuples(index=False):
        ws.append(list(r))

    team_col = df.columns.get_loc("team_name") + 1 if "team_name" in df.columns else None
    current_team = None
    fill_idx = 0

    for row_i in range(2, ws.max_row + 1):
        if team_col:
            team_val = ws.cell(row=row_i, column=team_col).value
            if team_val != current_team:
                current_team = team_val
                fill_idx = 1 - fill_idx
        for col_i in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_i, column=col_i)
            cell.fill = fills[fill_idx]
            cell.alignment = align

    # Autofit widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------- UI Controls -----------------------------

st.markdown('<div class="card">', unsafe_allow_html=True)

c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.2, 1.2])

with c1:
    base_url = st.text_input("Base URL", value=BASE_DEFAULT)
    seed_path = st.text_input("Seed path", value=SEED_PATH_DEFAULT)

with c2:
    qc_only = st.toggle("Québec only (province=QC)", value=False)
    city_contains = st.text_input("City filter (optional)", placeholder="e.g., Montréal, Quebec, Laval")

with c3:
    polite_delay = st.slider("Polite delay (seconds)", 0.0, 1.5, 0.25, 0.05)
    max_profiles = st.number_input("Max profiles (0 = no limit)", min_value=0, max_value=50000, value=0, step=100)

with c4:
    deep_crawl = st.toggle(
        "Deep crawl advisor pages (find more links)", value=False,
        help="Usually not needed. /advisor.html already contains the full list."
    )
    crawl_page_limit = st.number_input("Crawl page limit", min_value=10, max_value=5000, value=250, step=50, disabled=not deep_crawl)
    include_profile_url = st.toggle("Include profile URL column", value=False)
    include_address_hint = st.toggle("Include address hint column", value=False)
    include_city = st.toggle("Include city column", value=True)
    do_excel = st.toggle("Also generate Excel (.xlsx)", value=False, disabled=not OPENPYXL_OK)

debug_show_samples = st.toggle("Debug: show province fill samples", value=False)

run = st.button("Run Extraction", use_container_width=True)

st.markdown('</div>', unsafe_allow_html=True)

st.markdown(
    '<div class="small-muted">If province is missing on some profile pages, the app falls back to the directory page location.</div>',
    unsafe_allow_html=True,
)

st.divider()

if not run:
    st.info("Set your options above, then click **Run Extraction**.")
    st.stop()


# ----------------------------- Main -----------------------------

directory_url = urljoin(base_url, seed_path)
session = requests.Session()

status = st.empty()
progress = st.progress(0)

metrics = st.columns(4)
metrics[0].metric("Profile links", "0")
metrics[1].metric("Processed", "0")
metrics[2].metric("Kept", "0")
metrics[3].metric("Errors", "0")

errors = 0

# Load seed directory page
try:
    status.info(f"Loading seed page: {directory_url}")
    seed_html = safe_get(session, directory_url, delay_s=polite_delay)
except Exception as e:
    st.error(f"Failed to load seed page.\n\n{e}")
    st.stop()

# Build reliable city/province lookup from /advisor.html
dir_lookup = build_directory_location_lookup(seed_html, base_url)

# Collect advisor profile URLs from seed
advisor_urls = set(extract_advisor_urls_from_html(seed_html, base_url))

# Optional deep crawl
if deep_crawl:
    status.info("Deep crawling to discover more advisor profile links...")
    seen_pages = set()
    q = deque([directory_url])
    internal_pages_checked = 0

    while q and internal_pages_checked < int(crawl_page_limit):
        page = q.popleft()
        if page in seen_pages:
            continue
        seen_pages.add(page)

        try:
            html = safe_get(session, page, delay_s=polite_delay)
        except Exception:
            errors += 1
            continue

        advisor_urls.update(extract_advisor_urls_from_html(html, base_url))

        for nxt in extract_internal_html_pages_from_html(html, page, base_url):
            if nxt not in seen_pages:
                q.append(nxt)

        internal_pages_checked += 1

advisor_urls = sorted(advisor_urls)
metrics[0].metric("Profile links", f"{len(advisor_urls)}")

if not advisor_urls:
    st.warning("No advisor profile links found.")
    st.stop()

# Apply max profiles limit
if max_profiles and int(max_profiles) > 0:
    advisor_urls = advisor_urls[: int(max_profiles)]

total = len(advisor_urls)
rows = []
kept = 0

sample_rows = []

for i, url in enumerate(advisor_urls, start=1):
    try:
        status.info(f"Fetching {i}/{total}: {url}")
        html = safe_get(session, url, delay_s=polite_delay)
        row = parse_advisor_page(html, url=url, base=base_url, dir_lookup=dir_lookup)

        # Province filter
        if qc_only and row.get("province", "") != "QC":
            continue

        # City filter (best-effort: uses directory-derived city + address_hint)
        if city_contains.strip():
            target = city_contains.strip().lower()
            hay = f"{row.get('city','')} {row.get('address_hint','')}".lower()
            if target not in hay:
                continue

        rows.append(row)
        kept += 1

        if debug_show_samples and len(sample_rows) < 10:
            sample_rows.append({
                "name": row.get("name"),
                "province": row.get("province"),
                "city": row.get("city"),
                "address_hint": row.get("address_hint"),
                "profile_url": row.get("profile_url"),
            })

    except Exception:
        errors += 1

    progress.progress(int((i / total) * 100))
    metrics[1].metric("Processed", f"{i}")
    metrics[2].metric("Kept", f"{kept}")
    metrics[3].metric("Errors", f"{errors}")

status.success("Extraction complete.")

if not rows:
    st.warning("No advisors matched your filters.")
    st.stop()

df = pd.DataFrame(rows)
df = dedupe_rows(df)

# Output columns
out_cols = ["name", "email", "phone", "team_name", "province"]
if include_city:
    out_cols.insert(out_cols.index("province"), "city")
if include_address_hint:
    out_cols.append("address_hint")
if include_profile_url:
    out_cols.append("profile_url")

df_out = df[out_cols].copy()

# Show debug samples
if debug_show_samples:
    st.subheader("Debug samples (first 10 kept)")
    st.dataframe(pd.DataFrame(sample_rows), use_container_width=True, height=280)

# Preview
st.subheader("Preview")
st.dataframe(df_out, use_container_width=True, height=560)

# Download CSV
csv_bytes = df_out.to_csv(index=False).encode("utf-8")
st.download_button(
    "Download CSV",
    data=csv_bytes,
    file_name="nbfwm_advisors.csv",
    mime="text/csv",
    use_container_width=True,
)

# Download Excel
if do_excel and OPENPYXL_OK:
    try:
        xlsx_bytes = df_to_excel_bytes(df_out)
        st.download_button(
            "Download Excel (.xlsx)",
            data=xlsx_bytes,
            file_name="nbfwm_advisors.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as e:
        st.warning(f"Excel export failed: {e}")

# Quick sanity stat
blank_prov = int((df_out["province"].fillna("") == "").sum()) if "province" in df_out.columns else 0
st.caption(f"Province blanks: {blank_prov} / {len(df_out)}")

with st.expander("Notes / troubleshooting"):
    st.write(
        "- Province is extracted from the advisor page address when available.\n"
        "- If missing there, it falls back to the location shown on /advisor.html (City, Province).\n"
        "- If you redeployed and still see blanks, reboot the app and clear cache on Streamlit Cloud."
    )
```

If you want, paste 2–3 example profile URLs where you’re seeing blanks and I’ll tighten the parser even more — but this directory-fallback approach usually fixes it across the board.
